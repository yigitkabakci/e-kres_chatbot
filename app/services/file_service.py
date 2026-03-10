from __future__ import annotations

import csv
import json
import logging
import re
from pathlib import Path
from typing import Any

from app.schemas.chat import DailyMenu, DailySchedule, ScheduleItem
from app.services.base_service import BaseTool
from app.services.pdf_service import extract_text_from_pdf

logger = logging.getLogger(__name__)

DATA_DIR = Path(__file__).resolve().parents[2] / "data"
IMPORT_DIR = DATA_DIR / "imports"


class FileService:
    def __init__(self) -> None:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        IMPORT_DIR.mkdir(parents=True, exist_ok=True)
        logger.info("FileService hazir, data dizini: %s", DATA_DIR)

    def save_upload(self, filename: str, content: bytes) -> Path:
        safe_name = re.sub(r"[^a-zA-Z0-9._-]", "_", Path(filename).name)
        target = IMPORT_DIR / safe_name
        target.write_bytes(content)
        return target

    def read_excel(self, filename: str, sheet_name: str | int = 0) -> list[dict[str, Any]]:
        filepath = DATA_DIR / filename
        if not filepath.exists():
            logger.warning("Excel dosyasi bulunamadi: %s", filepath)
            return []
        return self.read_tabular_file(filepath, sheet_name=sheet_name)

    def read_tabular_file(self, filepath: Path, sheet_name: str | int = 0) -> list[dict[str, Any]]:
        suffix = filepath.suffix.lower()
        if suffix == ".csv":
            with filepath.open("r", encoding="utf-8-sig", newline="") as file:
                return list(csv.DictReader(file))
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            try:
                from openpyxl import load_workbook
            except ImportError:
                logger.error("openpyxl yuklu degil, Excel dosyasi okunamiyor.")
                return []
            workbook = load_workbook(filepath, data_only=True)
            sheet = workbook[workbook.sheetnames[sheet_name if isinstance(sheet_name, int) else 0]] if isinstance(sheet_name, int) else workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
            records: list[dict[str, Any]] = []
            for row in rows[1:]:
                record = {headers[index]: row[index] for index in range(min(len(headers), len(row))) if headers[index]}
                if any(value not in (None, "") for value in record.values()):
                    records.append(record)
            return records
        logger.warning("Desteklenmeyen tablo dosyasi: %s", filepath)
        return []

    def read_json(self, filename: str) -> Any:
        filepath = DATA_DIR / filename
        if not filepath.exists():
            logger.warning("JSON dosyasi bulunamadi: %s", filepath)
            return {}
        try:
            with filepath.open("r", encoding="utf-8") as file:
                return json.load(file)
        except Exception as exc:
            logger.error("JSON okuma hatasi (%s): %s", filename, exc)
            return {}

    def list_files(self) -> list[str]:
        if not DATA_DIR.exists():
            return []
        return [item.name for item in DATA_DIR.iterdir() if item.is_file()]

    def save_json(self, filename: str, data: Any) -> bool:
        filepath = DATA_DIR / filename
        try:
            with filepath.open("w", encoding="utf-8") as file:
                json.dump(data, file, ensure_ascii=False, indent=2)
            return True
        except Exception as exc:
            logger.error("JSON kaydetme hatasi (%s): %s", filename, exc)
            return False

    def parse_admin_data_file(self, filepath: Path, section: str) -> dict[str, Any]:
        section_key = section.lower()
        suffix = filepath.suffix.lower()
        if suffix == ".pdf":
            text, metric = extract_text_from_pdf(str(filepath))
            if not text:
                raise ValueError("PDF dosyasi parse edilemedi.")
            if section_key == "meals":
                meals = self._parse_meals_from_text(text)
                return {"meals": meals, "text": text, "metric": metric}
            schedules, daily_flow = self._parse_schedule_from_text(text)
            return {"schedules": schedules, "daily_flow": daily_flow, "text": text, "metric": metric}

        rows = self.read_tabular_file(filepath)
        if not rows:
            raise ValueError("Dosyada parse edilebilir veri bulunamadi.")
        if section_key == "meals":
            meals = self._parse_meals_from_rows(rows)
            return {"meals": meals, "text": json.dumps(meals, ensure_ascii=False), "metric": len(rows)}
        schedules, daily_flow = self._parse_schedule_from_rows(rows)
        return {
            "schedules": schedules,
            "daily_flow": daily_flow,
            "text": json.dumps({"schedules": schedules, "daily_flow": daily_flow}, ensure_ascii=False),
            "metric": len(rows),
        }

    @staticmethod
    def _normalize_key(key: Any) -> str:
        return str(key or "").strip().lower().replace("ı", "i")

    @staticmethod
    def _split_list(value: Any) -> list[str]:
        if value is None:
            return []
        if isinstance(value, list):
            return [str(item).strip() for item in value if str(item).strip()]
        return [item.strip() for item in str(value).split(",") if item.strip()]

    def _parse_meals_from_rows(self, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        meals: list[dict[str, Any]] = []
        for row in rows:
            normalized = {self._normalize_key(key): value for key, value in row.items()}
            meal = DailyMenu.model_validate({
                "tarih": normalized.get("tarih") or normalized.get("date"),
                "kahvalti": self._split_list(normalized.get("kahvalti")),
                "ogle": self._split_list(normalized.get("ogle")),
                "ikindi": self._split_list(normalized.get("ikindi")),
                "ara_ogun": self._split_list(normalized.get("ara ogun") or normalized.get("ara_ogun")),
                "aciklama": normalized.get("aciklama"),
            })
            meals.append(meal.model_dump(mode="json"))
        return meals

    def _parse_schedule_from_rows(self, rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
        grouped: dict[str, list[dict[str, str]]] = {}
        daily_flow: list[dict[str, str]] = []
        for row in rows:
            normalized = {self._normalize_key(key): value for key, value in row.items()}
            gun = str(normalized.get("gun") or "").strip()
            saat = str(normalized.get("saat") or normalized.get("time") or "").strip()
            etkinlik = str(normalized.get("etkinlik") or normalized.get("activity") or "").strip()
            akis = str(normalized.get("akis") or normalized.get("daily_flow") or normalized.get("akis") or "").strip()
            if gun and saat and etkinlik:
                grouped.setdefault(gun, []).append({"saat": saat, "etkinlik": etkinlik})
            elif saat and (akis or etkinlik):
                daily_flow.append({"time": saat, "activity": akis or etkinlik})
        schedules = [DailySchedule(gun=gun, dersler=[ScheduleItem(**lesson) for lesson in lessons]).model_dump(mode="json") for gun, lessons in grouped.items()]
        return schedules, daily_flow

    def _parse_meals_from_text(self, text: str) -> list[dict[str, Any]]:
        blocks = [block.strip() for block in re.split(r"\n\s*\n", text) if block.strip()]
        meals: list[dict[str, Any]] = []
        current_date = None
        for block in blocks:
            date_match = re.search(r"(\d{4}-\d{2}-\d{2})", block)
            if date_match:
                current_date = date_match.group(1)
            if not current_date:
                continue
            kahvalti = self._extract_labeled_list(block, ["Kahvalti", "Kahvaltı"])
            ogle = self._extract_labeled_list(block, ["Ogle", "Öğle"])
            ikindi = self._extract_labeled_list(block, ["Ikindi", "İkindi"])
            ara_ogun = self._extract_labeled_list(block, ["Ara Ogun", "Ara Öğün"])
            if kahvalti or ogle or ikindi or ara_ogun:
                meals.append(DailyMenu.model_validate({
                    "tarih": current_date,
                    "kahvalti": kahvalti,
                    "ogle": ogle,
                    "ikindi": ikindi,
                    "ara_ogun": ara_ogun,
                    "aciklama": None,
                }).model_dump(mode="json"))
        if not meals:
            raise ValueError("PDF icinden yemek verisi cikartilamadi.")
        return meals

    def _parse_schedule_from_text(self, text: str) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
        grouped: dict[str, list[dict[str, str]]] = {}
        daily_flow: list[dict[str, str]] = []
        current_day = None
        for raw_line in text.splitlines():
            line = raw_line.strip(" -\t")
            if not line:
                continue
            if line in {"Pazartesi", "Sali", "Salı", "Carsamba", "Çarşamba", "Persembe", "Perşembe", "Cuma"}:
                current_day = line.replace("Salı", "Sali").replace("Çarşamba", "Carsamba").replace("Perşembe", "Persembe")
                grouped.setdefault(current_day, [])
                continue
            lesson_match = re.match(r"(?P<time>\d{2}:\d{2}\s*-\s*\d{2}:\d{2})\s*[:|-]?\s*(?P<activity>.+)", line)
            if lesson_match and current_day:
                grouped[current_day].append({"saat": lesson_match.group("time").replace(" ", ""), "etkinlik": lesson_match.group("activity").strip()})
                continue
            flow_match = re.match(r"(?P<time>\d{2}:\d{2})\s*[:|-]?\s*(?P<activity>.+)", line)
            if flow_match:
                daily_flow.append({"time": flow_match.group("time"), "activity": flow_match.group("activity").strip()})
        schedules = [DailySchedule(gun=gun, dersler=[ScheduleItem(**lesson) for lesson in lessons]).model_dump(mode="json") for gun, lessons in grouped.items() if lessons]
        if not schedules and not daily_flow:
            raise ValueError("PDF icinden ders programi veya gunluk akis verisi cikartilamadi.")
        return schedules, daily_flow

    @staticmethod
    def _extract_labeled_list(text: str, labels: list[str]) -> list[str]:
        for label in labels:
            match = re.search(label + r"\s*[:|-]\s*(.+)", text, flags=re.IGNORECASE)
            if match:
                return [item.strip() for item in re.split(r",|/", match.group(1)) if item.strip()]
        return []


class FileQueryTool(BaseTool):
    @property
    def name(self) -> str:
        return "file_query"

    @property
    def description(self) -> str:
        return "data/ klasorundeki Excel ve JSON dosyalarini okur. Ogrenci listesi, etkinlik takvimi gibi verilere erisir."

    def __init__(self) -> None:
        self._fs = FileService()

    async def run(self, query: str, **kwargs: Any) -> dict[str, Any]:
        files = self._fs.list_files()
        if not files:
            return {
                "type": "file_query",
                "message": "data/ klasorunde henuz dosya yok. Ogretmen Excel, CSV, PDF veya JSON dosyasi yukleyebilir.",
            }
        return {
            "type": "file_query",
            "data": {"files": files},
            "message": "Mevcut dosyalar: " + ", ".join(files),
        }
